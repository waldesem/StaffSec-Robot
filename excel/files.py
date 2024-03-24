import os
import logging
import shutil
from datetime import datetime, date

from openpyxl import load_workbook

from config import Config
from parsers.excelparser import screen_excel
from parsers.jsonparser import screen_json
from database.dbase import db_iquiry_data, db_main_data
from action.actions import normalize_name


"""Parses the main Excel file to get today's records, screens them, 
archives processed records, and saves results to the database.

- Loads the main Excel file 
- Iterates over today's records
    - Gets key data for the record
    - Finds the matching subdirectory
    - Screens Excel/JSON files in the subdirectory
    - Archives the subdirectory
    - Saves record data to the database
- Saves the updated main Excel file
"""
def parse_main(file):
    wb = load_workbook(file, keep_vba=True)
    ws = wb.worksheets[0]

    today_records = {}
    for i, cell in enumerate(ws["K10000":"K50000"], 10000):
        for c in cell:
            if (
                c.value
                and isinstance(c.value, datetime)
                and (c.value).date() == date.today()
            ):
                today_records.update({
                    str(i): {
                        "fullname": normalize_name(ws["B" + str(i)].value),
                        "birthday": 
                            ws["C" + str(i)].value.date()
                            if isinstance(ws["C" + str(i)].value, datetime)
                            else date.today(),
                        "id": ws['A' + str(i)].value,
                    }
                })

    subdirs = os.listdir(Config.WORK_DIR)
    for key, value in today_records.items():
        link = ""

        for sub in subdirs:
            if normalize_name(sub) == value["fullname"]:
                subdir_path = os.path.join(Config.WORK_DIR, sub)

                scan_subdir(subdir_path)

                link = os.path.join(
                            Config.ARCHIVE_DIR,
                            sub[0],
                            f"{sub} - {value['id']}"
                )
                if link:
                    ws["L" + key].hyperlink = link

                move_subdir(subdir_path, sub, link)

                break

        db_main_data(value['fullname'], value['birthday'], link)

    wb.save(file)
    wb.close()
    logging.info("Main file parsed")


"""Parse all files in a subdirectory.

This parses all files in the given subdirectory path that match certain 
naming patterns. Excel files starting with "Заключение" or "Результаты" 
are passed to screen_excel(). JSON files are passed to screen_json().

Args:
    subdir_path: The subdirectory path to parse files in.
"""
def scan_subdir(subdir_path):
    for file in os.listdir(subdir_path):
        if file.endswith("xlsm") or file.endswith("xlsx"):
            screen_excel(subdir_path, file)
        elif file.endswith("json"):
            screen_json(subdir_path, file)


"""Move a subdirectory after processing.

This moves the given subdirectory path to the archive directory if it does not already exist there. 
If the archive directory already contains a directory with the same name, 
it will log a message and not move it again. Any exceptions are also logged.

Args:
    subdir_path: The subdirectory path to move.
    sub: The name of the subdirectory.
    link: The path to the archive directory.
"""
def move_subdir(subdir_path, sub, link):
    try:
        if not os.path.isdir(link):
            shutil.move(subdir_path, link)
            logging.info(f"{sub} moved to {Config.ARCHIVE_DIR}")
        else:
            logging.info(f"{sub} already in {Config.ARCHIVE_DIR}")
    except Exception as e:
        logging.error(e)


"""Parse the inquiry worksheet in the info Excel file.

Iterates through the date column, checking for today's date. 
For each row with today's date, extracts the info, initiator, 
fullname, and birthday, and saves to the database.

After processing the whole worksheet, closes the workbook and logs 
that the info file was parsed.
"""
def parse_inquiry(file):
    wb = load_workbook(file, keep_vba=True)
    ws = wb.worksheets[0]
    for i, cell in enumerate(ws["G500":"G5000"], 500):
        for c in cell:
            if (
                c.value
                and isinstance(c.value, datetime)
                and (c.value).date() == date.today()
            ):
                info = ws[f"E{i}"].value
                initiator = ws[f"F{i}"].value
                fullname = normalize_name(ws[f"A{i}"].value)
                birthday = (
                    (ws[f"B{i}"].value).date()
                    if isinstance(ws[f"B{i}"].value, datetime)
                    else date.today()
                )
                db_iquiry_data(info, initiator, fullname, birthday)
    wb.close()
    logging.info("Info file parsed")
