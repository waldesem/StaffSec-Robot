import os
from datetime import date, datetime

from openpyxl import load_workbook

from database.dbase import excel_to_db
from action.actions import name_convert, get_item_id


"""Parses an Excel file to extract resume and other data.

Args:
  excel_path: Path to the directory containing the Excel file.
  excel_file: Name of the Excel file.

Returns:
  None. Data is inserted into the database.

This function loads the given Excel file, extracts resume and other data 
based on the sheet name and cell contents, converts names, gets item IDs,
and inserts the data into the database if valid.
"""
def screen_excel(excel_path, excel_file):
    workbook = load_workbook(os.path.join(excel_path, excel_file), keep_vba=True)
    worksheet = workbook.worksheets[0]
    person = {}

    if excel_file.startswith("Заключение"):
        if len(workbook.sheetnames) > 1:
            sheet = workbook.worksheets[1]

            if (
                str(sheet["K1"].value) == "ФИО"
                and sheet["K3"].value
                and sheet["L3"].value
            ):
                person.update({"resume": get_resume(sheet)})

        if worksheet["C6"].value and worksheet["C8"].value:
            person.update({"resume": get_conclusion_resume(worksheet)})

        if worksheet["C23"].value:
            person.update({"check": get_check(worksheet)})
    else:
        person.update({"resume": get_robot_resume(worksheet)})
        person.update({"robot": get_robot(worksheet)})

    workbook.close()

    if "resume" in person:
        if (
            person["resume"]["fullname"]
            and person["resume"]["birthday"] != date.today()
        ):
            excel_to_db(person)


"""Parses resume data from the given sheet.

Args:
  sheet: The sheet containing resume data.

Returns: 
  A dict with the parsed resume fields.
"""
def get_resume(sheet):
    return {
        "fullname": name_convert(str(sheet["K3"].value)),
        "birthday": (
            (sheet["L3"].value).date()
            if isinstance(sheet["L3"].value, datetime)
            else (
                datetime.strptime(sheet["L3"].value, "%d.%m.%Y").date()
                if sheet["L3"].value
                else date.today()
            )
        ),
        "birthplace": str(sheet["M3"].value).strip(),
        "country": str(sheet["T3"].value).strip(),
        "snils": str(sheet["U3"].value).strip(),
        "inn": str(sheet["V3"].value).strip(),
    }


"""Parses resume conclusion data from the given sheet.

Args:
  sheet: The sheet containing resume conclusion data. 

Returns:
  A dict with the parsed resume conclusion fields.
"""
def get_conclusion_resume(sheet):
    return {
        "fullname": name_convert(sheet["C6"].value),
        "birthday": (
            (sheet["C8"].value).date()
            if isinstance(sheet["C8"].value, datetime)
            else (
                datetime.strptime(sheet["C8"].value, "%d.%m.%Y").date()
                if sheet["C8"].value
                else date.today()
            )
        ),
    }


"""Parses robot resume data from the given sheet.

Args:
  sheet: The sheet containing robot resume data.

Returns:
  A dict with the parsed robot resume fields.
"""
def get_robot_resume(sheet):
    return {
        "fullname": name_convert(sheet["B4"].value),
        "birthday": (
            (sheet["B5"].value).date()
            if isinstance(sheet["B5"].value, datetime)
            else (
                datetime.strptime(sheet["B5"].value, "%d.%m.%Y").date()
                if sheet["B5"].value
                else date.today()
            )
        ),
    }


"""Parses check data from the given sheet.

Args:
  sheet: The sheet containing check data.

Returns:
  A dict with the parsed check fields.
"""
def get_check(sheet):
    return {
        "workplace": f"{sheet['C11'].value} - {sheet['D11'].value}; {sheet['C12'].value} - "
        f"{sheet['D12'].value}; {sheet['C13'].value} - {sheet['D13'].value}",
        "cronos": f"{sheet['B14'].value}: {sheet['C14'].value}; {sheet['B15'].value}: "
        f"{sheet['C15'].value}",
        "cros": sheet["C16"].value,
        "document": sheet["C17"].value,
        "debt": sheet["C18"].value,
        "bankruptcy": sheet["C19"].value,
        "bki": sheet["C20"].value,
        "affilation": sheet["C21"].value,
        "internet": sheet["C22"].value,
        "pfo": (
            False
            if sheet["C26"].value or str(sheet["C26"].value).lower() == "не назначалось"
            else True
        ),
        "addition": sheet["C28"].value,
        "conclusion": get_item_id(
            "conclusions", "conclusion", sheet["C23"].value
        ),
        "deadline": datetime.now(),
        "officer": sheet["C25"].value,
    }


"""Parses robot check data from the given sheet.

Args:
  sheet: The sheet containing robot check data. 

Returns:
  A dict with the parsed robot check fields.
"""
def get_robot(sheet):
    return {
        "employee": sheet["B27"].value,
        "terrorist": sheet["B17"].value,
        "inn": sheet["B18"].value,
        "bankruptcy": f"{sheet['B20'].value}, {sheet['B21'].value}, {sheet['B22'].value}",
        "mvd": sheet["B23"].value,
        "courts": sheet["B24"].value,
        "bki": sheet["B25"].value,
        "deadline": datetime.now(),
    }
