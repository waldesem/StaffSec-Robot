import os
import shutil
from datetime import datetime, date

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from config import Config
from models.model import engine,Person, Inquiry, Check
from .excelparser import excel_to_db
from .jsonparser import json_to_db


def parse_main():
    wb = openpyxl.load_workbook(Config.MAIN_FILE, keep_vba=True, read_only=False)
    ws = wb.worksheets[0]
    num_row = range_row(ws['K5000':'K25000'])

    if len(num_row):
        # list of directories with candidates names
        fio = [ws['B' + str(i)].value.strip().lower() for i in num_row]
        subdir = [sub for sub in os.listdir(Config.WORK_DIR) if sub.lower().strip() in fio]

        if len(subdir):
            # получаем список путей к файлам Заключений
            excel_path = list(filter(None, []))
            json_path = list(filter(None, []))

            for f in subdir:
                subdir_path = os.path.join(Config.WORK_DIR[0:-1], f)
                for file in os.listdir(subdir_path):
                    if (file.startswith("Заключение") or file.startswith("Результат")) and (file.endswith("xlsm") or file.endswith("xlsx")):
                        excel_path.append(os.path.join(Config.WORK_DIR, subdir_path, file))
                    elif file.endswith("json"):
                        json_path.append(os.path.join(Config.WORK_DIR, subdir_path, file))

            # parse files and send info to database            
            if len(excel_path):
                excel_to_db(excel_path)  
            if len(json_path):
                json_to_db(json_path)

            # create url and move folders to archive
            for n in num_row:
                for sub in subdir:
                    if str(ws['B' + str(n)].value.strip().lower()) == sub.strip().lower():
                        sbd = ws['B' + str(n)].value.strip()
                        lnk = os.path.join(Config.ARCHIVE_DIR, sbd[0][0], f"{sbd} - {ws['A' + str(n)].value}")
                        ws['L' + str(n)].hyperlink = str(lnk)  # записывает в книгу
                        shutil.move(os.path.join(Config.WORK_DIR, sbd), lnk)
                        
            screen_registry_data(ws, num_row)
        wb.save(Config.MAIN_FILE)
    else:
        wb.close()


def parse_info():
    wb = openpyxl.load_workbook(Config.INFO_FILE, keep_vba=True, read_only=False)
    ws = wb.worksheets[0]
    num_row = range_row(ws['G1':'G3000'])
    if len(num_row):
        screen_iquiry_data(ws, num_row)
    wb.close()


def range_row(sheet):
    row_num = []
    for cell in sheet:
        for c in cell:
            if isinstance(c.value, datetime):  # check format of date
                if c.value.strftime('%Y-%m-%d') == date.today().strftime('%Y-%m-%d'):
                    row_num.append(c.row)
            elif str(c.value).strip() == date.today().strftime('%d.%m.%Y'):
                row_num.append(c.row)
    return row_num

        
def screen_iquiry_data(sheet, num_row): 
    for num in num_row:
        chart = {'info': sheet[f'E{num}'].value,
                 'initiator': sheet[f'F{num}'].value,
                 'deadline': datetime.strftime(datetime.strptime(str(sheet[f'G{num}'].value).\
                             strip(), '%d.%m.%Y'), '%Y-%m-%d'),
                 'fullname': sheet['A' + str(num)].value,
                 'birthday': sheet['B' + str(num)].value}
        with Session(engine) as sess:
            person = sess.execute(
                select(Person)
                .filter_by(fullname=chart['fullname'], 
                           birthday=chart['birthday'])
                ).scalar_one_or_none()
            if not person:
                person = Person(Person(fullname=chart['fullname'],
                                birthday=chart['birthday']))
                sess.add(person)
                sess.flush()
            sess.add(Inquiry(info = chart['info'],
                                initiator = chart['initiator'],
                                deadline = chart['deadline'],
                                person_id = person.id))
            sess.commit()


def screen_registry_data(sheet, num_row):
    for num in num_row:
        chart = {'fullname': sheet['A' + str(num)].value,
                 'birthday': sheet['B' + str(num)].value,
                 'decision': sheet[f'J{num}'].value,
                 'deadline': datetime.strftime(datetime.strptime(str(sheet[f'K{num}'].value).\
                             strip(), '%d.%m.%Y'), '%Y-%m-%d'),
                 'url': sheet[f'L{num}'].value}
        with Session(engine) as sess:
            person = sess.execute(
                select(Person)
                .filter_by(fullname=chart['fullname'], 
                           birthday=chart['birthday'])
                ).scalar_one_or_none()
            if not person:
                person = Person(Person(fullname=chart['fullname'],
                                birthday=chart['birthday']))
                sess.add(person)
                sess.flush()
            sess.add(Check(decision = chart['decision'],
                           deadline = chart['deadline'],
                           url = chart['url'],
                           person_id = person.id))
            sess.commit()