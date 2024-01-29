import os
from datetime import date, datetime

import aiosqlite
from openpyxl import load_workbook

from config import Config


async def excel_to_db(excel_path, excel_file):
    excel = await screen_excel(excel_path, excel_file)
    if excel.get("resume"):
        excel["resume"] | {"category_id": 1, "status_id": 9, "region_id": 1}

        async with aiosqlite.connect(Config.DATABASE_URI) as db:
            async with db.execute(
                "SELECT * FROM persons WHERE fullname = ? AND birthday = ?",
                (excel["resume"]["fullname"], excel["resume"]["birthday"]),
            ) as cursor:
                result = await cursor.fetchone()
                person_id = result[0] if result else None

            if not person_id:
                await db.execute(
                    f"INSERT INTO persons ({','.join(excel['resume'].keys())}), created "
                    f"VALUES ({','.join(['?'] * len(excel['resume'].values()))}), ?",
                    tuple(excel["resume"].values()) + (datetime.now(),),
                )
                person_id = cursor.lastrowid
            else:
                await db.execute(
                    f"UPDATE persons SET {'=?,'.join(excel['resume'].keys())}=?, updated=? "
                    f"WHERE id = {person_id}",
                    tuple(excel["resume"].values()) + (datetime.now(),),
                )

            if excel.get("check"):
                await db.execute(
                    f"INSERT INTO checks ({','.join(excel['check'].keys())}, person_id) "
                    f"VALUES ({','.join(['?'] * len(excel['check'].values()))}, ?)",
                    tuple(excel["check"].values()) + (person_id,),
                )
            elif excel.get("robot"):
                await db.execute(
                    f"INSERT INTO robots ({','.join(excel['robot'].keys())}, person_id) "
                    f"VALUES ({','.join(['?'] * len(excel['robot'].values()))}, ?)",
                    tuple(excel["robot"].values()) + (person_id,),
                )

            await db.commit()


async def screen_excel(excel_path, excel_file):
    workbook = load_workbook(os.path.join(excel_path, excel_file), keep_vba=True)
    worksheet = workbook.worksheets[0]
    person = {}

    if excel_file.startswith("Заключение"):
        if len(workbook.sheetnames) > 1:
            sheet = workbook.worksheets[1]
            if str(sheet["K1"].value) == "ФИО":
                person.update({"resume": await get_resume(sheet)})
        person.update({"resume": await get_conclusion_resume(worksheet)})
        person.update({"check": await get_check(worksheet)})
    else:
        person.update({"resume": await get_robot_resume(worksheet)})
        person.update({"robot": await get_robot(worksheet)})

    workbook.close()
    return person


async def get_resume(sheet):
    resume = {
        "fullname": fullname_parser(str(sheet["K3"].value)),
        "birthday": datetime.strptime(sheet["L3"].value, "%d.%m.%Y").date()
        if sheet["L3"].value
        else date.today(),
        "birthplace": str(sheet["M3"].value).strip(),
        "country": str(sheet["T3"].value).strip(),
        "snils": str(sheet["U3"].value).strip(),
        "inn": str(sheet["V3"].value).strip(),
    }
    return resume


async def get_conclusion_resume(sheet):
    resumes = {
        "fullname": fullname_parser(sheet["C6"].value),
        "birthday": (sheet["C8"].value).date()
        if isinstance(sheet["L3"].value, datetime)
        else date.today(),
    }
    return resumes


async def get_robot_resume(sheet):
    resumes = {
        "fullname": fullname_parser(sheet["B4"].value),
        "birthday": datetime.strptime(sheet["B5"].value, "%d.%m.%Y").date()
        if sheet["B5"].value
        else date.today(),
    }
    return resumes


async def get_check(sheet):
    checks = {
        "workplace": f"{sheet['C11'].value} - {sheet['D11'].value}; {sheet['C12'].value} - "
        f"{sheet['D12'].value}; {sheet['C13'].value} - {sheet['D13'].value}",
        "cronos": f"{sheet['B14'].value}: {sheet['C14'].value}; {sheet['B15'].value}: "
        f"{sheet['C15'].value}",
        "cros": sheet["C16"].value,
        "document": sheet["C17"].value,
        "debt": sheet["C18"].value,
        "bankruptcy": sheet["C19"].value,
        "bki": sheet["C20"].value,
        "affilation": sheet["C21"].value,
        "internet": sheet["C22"].value,
        "pfo": True if sheet["C26"].value else False,
        "addition": sheet["C28"].value,
        "conclusion": await get_conclusion_id(sheet["C23"].value),
        "deadline": datetime.now(),
        "officer": sheet["C25"].value,
    }
    return checks


async def get_robot(sheet):
    robot = {
        "employee": sheet["B27"].value,
        "terrorist": sheet["B17"].value,
        "inn": sheet["B18"].value,
        "bankruptcy": f"{sheet['B20'].value}, {sheet['B21'].value}, {sheet['B22'].value}",
        "mvd": sheet["B23"].value,
        "courts": sheet["B24"].value,
        "bki": sheet["B25"].value,
        "deadline": datetime.now(),
    }
    return robot


async def get_conclusion_id(name):
    async with aiosqlite.connect(Config.DATABASE_URI) as conn:
        async with conn.execute(
            "SELECT * FROM conclusions WHERE LOWER (conclusion) = ?", (name.lower(),)
        ) as cursor:
            result = await cursor.fetchone()
            return result[0] if result else 1


def fullname_parser(fullname: str) -> str:
    return " ".join(filter(None, map(str.strip, fullname.split()))).upper()
