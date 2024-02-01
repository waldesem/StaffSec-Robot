import os
import shutil
from datetime import datetime, date
import logging
import asyncio

import aiosqlite
from openpyxl import load_workbook

from config import Config
from excelparser import excel_to_db, fullname_parser, get_conclusion_id
from jsonparser import json_to_db


logging.basicConfig(
    filename=Config.LOG_FILE, format="%(asctime)s - %(message)s", level=logging.INFO
)


async def main():
    now = datetime.now()

    main_file_date = date.fromtimestamp(os.path.getmtime(Config.MAIN_FILE))
    info_file_date = date.fromtimestamp(os.path.getmtime(Config.INFO_FILE))

    tasks = [
        archive_db(main_file_date, info_file_date),
        archive_main(main_file_date),
        archive_info(info_file_date),
    ]
    await asyncio.gather(*tasks)

    logging.info(f"Script execution time: {datetime.now() - now}")


async def archive_db(main_file_date, info_file_date):
    if date.today() in [main_file_date, info_file_date]:
        shutil.copy(os.path.join(Config.BASE_PATH, "persons.db"), Config.ARCHIVE_DIR)
        logging.info(f"persons.db copied to {Config.ARCHIVE_DIR}")
    else:
        logging.info(f"persons.db not changed")


async def archive_main(main_file_date):
    if date.today() == main_file_date:
        shutil.copy(Config.MAIN_FILE, Config.ARCHIVE_DIR)
        logging.info(f"Main file copied to {Config.ARCHIVE_DIR}")
        await parse_main()
    else:
        logging.info(f"Main file not changed")


async def archive_info(info_file_date):
    if date.today() == info_file_date:
        shutil.copy(Config.INFO_FILE, Config.ARCHIVE_DIR)
        logging.info(f"Info file copied to {Config.ARCHIVE_DIR}")
        await parse_inquiry()
    else:
        logging.info(f"Info file not changed")


async def parse_main():
    wb = load_workbook(Config.MAIN_FILE, keep_vba=True)
    ws = wb.worksheets[0]

    subdirs = os.listdir(Config.WORK_DIR)
    for i, cell in enumerate(ws["K10000":"K50000"], 10000):
        for c in cell:
            if (
                c.value
                and isinstance(c.value, datetime)
                and (c.value).date() == date.today()
            ):
                fullname = fullname_parser(ws["B" + str(i)].value)
                birthday = (
                    (ws["C" + str(i)].value).date()
                    if isinstance(ws["C" + str(i)].value, datetime)
                    else date.today()
                )
                decision = ws[f"J{i}"].value
                lnk = ""

                for sub in subdirs:
                    if fullname_parser(sub) == fullname:
                        subdir_path = os.path.join(Config.WORK_DIR, sub)

                        for file in os.listdir(subdir_path):
                            if (
                                file.startswith("Заключение")
                                or file.startswith("Результаты")
                            ) and (file.endswith("xlsm") or file.endswith("xlsx")):
                                await excel_to_db(subdir_path, file)
                            elif file.endswith("json"):
                                await json_to_db(subdir_path, file)

                        lnk = os.path.join(
                            Config.ARCHIVE_DIR,
                            sub[0],
                            f"{sub} - {ws['A' + str(i)].value}",
                        )
                        ws["L" + str(i)].hyperlink = str(lnk)

                        try:
                            shutil.move(subdir_path, lnk)
                            logging.info(f"{sub} moved to {Config.ARCHIVE_DIR}")
                        except Exception as e:
                            logging.error(e)

                        break

                await db_main_data(fullname, birthday, decision, lnk)

    wb.save(Config.MAIN_FILE)
    wb.close()
    logging.info("Main file parsed")


async def db_main_data(fullname, birthday, decision, url):
    async with aiosqlite.connect(Config.DATABASE_URI) as db:
        async with db.execute(
            "SELECT * FROM persons WHERE fullname = ? AND birthday = ?",
            (fullname, birthday),
        ) as cursor:
            result = await cursor.fetchone()
            if not result:
                await db.execute(
                    f"INSERT INTO persons (fullname, birthday, path, created, category_id, region_id, status_id) "
                    f"VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (fullname, birthday, url, datetime.now(), 1, 1, 9),
                )

            else:
                await db.execute(
                    "UPDATE persons SET path = ?, updated = ? WHERE id = ?",
                    (url, datetime.now(), result[0]),
                )
            
            await db.execute(
                "INSERT INTO checks (conclusion, deadline, person_id) VALUES (?, ?, ?)",
                (
                    await get_conclusion_id(decision),
                    datetime.now(),
                    cursor.lastrowid,
                ),
            )

            await db.commit()


async def parse_inquiry():
    wb = load_workbook(Config.INFO_FILE, keep_vba=True)
    ws = wb.worksheets[0]
    for i, cell in enumerate(ws["G500":"G5000"], 500):
        for c in cell:
            if (
                c.value
                and isinstance(c.value, datetime)
                and (c.value).date() == date.today()
            ):
                info = ws[f"E{i}"].value
                initiator = ws[f"F{i}"].value
                fullname = fullname_parser(ws[f"A{i}"].value)
                birthday = (
                    (ws[f"B{i}"].value).date()
                    if isinstance(ws[f"B{i}"].value, datetime)
                    else date.today()
                )
                await db_iquiry_data(info, initiator, fullname, birthday)
    wb.close()
    logging.info("Info file parsed")


async def db_iquiry_data(info, initiator, fullname, birthday):
    async with aiosqlite.connect(Config.DATABASE_URI) as db:
        async with db.execute(
            "SELECT * FROM persons WHERE fullname = ? AND birthday = ?",
            (fullname, birthday),
        ) as cursor:
            result = await cursor.fetchone()

            if not result:
                await db.execute(
                    "INSERT INTO persons (fullname, birthday, created, category_id, region_id, status_id) \
                                VALUES (?, ?, ?, ?, ?, ?)",
                    (fullname, birthday, datetime.now(), 1, 1, 9),
                )
                result = [cursor.lastrowid]

            await db.execute(
                "INSERT INTO inquiries (info, initiator, deadline, person_id) \
                            VALUES (?, ?, ?, ?)",
                (info, initiator, datetime.now(), result[0]),
            )

            await db.commit()


if __name__ == "__main__":
    asyncio.run(main())
