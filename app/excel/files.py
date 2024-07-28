import logging
import os
import shutil

import openpyxl

from config import Config
from action.actions import normalize_name, check_date, check_birthday
from database.dbase import db_iquiry_data, db_main_data, excel_to_db, json_to_db
from parsers.excelparser import screen_excel
from parsers.jsonparser import screen_json
from parsers.xmlparser import screen_xml


def parse_registry(file):
    wb = openpyxl.load_workbook(file, keep_vba=True)
    ws = wb.worksheets[0]

    if file.endswith("Кандидаты.xlsm"):
        subdirs = [normalize_name(sub) for sub in os.listdir(Config.WORK_DIR)]
        for cell in ws["K2" : f"K{ws.max_row}"]:
            for c in cell:
                if check_date(c.value):
                    id = ws[f"A{c.row}"].value
                    fullname = normalize_name(ws[f"B{c.row}"].value)
                    birthday = check_birthday(ws[f"C{c.row}"].value)
                    link = ""
                    if fullname in subdirs:
                        link = os.path.join(
                            Config.ARCHIVE_DIR,
                            fullname[0],
                            f"{fullname} - {id}",
                        )
                        ws[f"L{c.row}"].hyperlink = link
                        parse_subdirs(fullname, link)
                    db_main_data(fullname, birthday, link)
        wb.save(file)
    else:
        for cell in ws["G2" : f"G{ws.max_row}"]:
            for c in cell:
                if check_date(c.value):
                    info = ws[f"E{c.row}"].value
                    initiator = ws[f"F{c.row}"].value
                    fullname = normalize_name(ws[f"A{c.row}"].value)
                    birthday = check_birthday(ws[f"B{c.row}"].value)
                    db_iquiry_data(info, initiator, fullname, birthday)

    wb.close()
    logging.info(f"{file} file parsed")


def parse_subdirs(fullname, link):
    subdir_path = os.path.join(Config.WORK_DIR, fullname)
    person = None
    json_string = None
    for file in os.listdir(subdir_path):
        if (file.endswith("xlsm") or file.endswith("xlsx")) and file.startswith(
            "Заключение"
        ):
            person = screen_excel(subdir_path, file)
        if file.endswith("json"):
            json_data = screen_json(subdir_path, file)
            json_to_db(json_data)
        if file.endswith("xml"):
            json_string = screen_xml(subdir_path, file)
    if (
        "resume" in person
        and "fullname" in person["resume"]
        and "birthday" in person["resume"]
    ):  
        if person.get("check"):
            person["check"]['addition'] = json_string
        excel_to_db(person)
    try:
        if not os.path.isdir(link):
            shutil.move(subdir_path, link)
            logging.info(f"{subdir_path} moved to archive")
        else:
            logging.info(f"{subdir_path} already in archive")
    except Exception as e:
        logging.error(e)
